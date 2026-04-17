import pandas as pd
import numpy as np
import matplotlib.pyplot as plt 
from openpyxl import workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from datetime import datetime

def excel_gen():

    filename = 'Week 2/Day 5/excel_gen_files/business.csv'
    img = 'Week 2/Day 5/excel_gen_files/sales_analytics.png'
    excel_name = 'Week 2/Day 5/excel_gen_files/business_report.xlsx'

    def load_business_data(filename):
        try:
            df = pd.read_csv(filename, sep = ',', na_values = ['NaN','nan','NAN','None','none','N/A','NA'], header = 0)
            df.replace(['nan','NAN','NaN','None','none','N/A','NA'], np.nan)
            df['date'] = pd.to_datetime(df['date'])
            df.sort_values('date', ascending = True)
            print("Business data from file:")
            print(df)
            return df

        except Exception as e:
            print(f"Unable to load business data: {e}")
    
    def create_summary_sheet(df):
        try:
           df['revenue'] = df['price'] * df['quantity']
           df['previous_revenue'] = df.groupby('category')['revenue'].shift(1)
           df['growth_rate'] = (((df['revenue'] - df['previous_revenue']) / df['previous_revenue']) * 100).round(2)
           
           idx = df.groupby('category')['revenue'].idxmax()
           top_map = df.loc[idx, ['category', 'salesperson']].set_index('category')['salesperson']
           df['category_top_performer'] = df['category'].map(top_map)
           
           df['cat_total_rv'] = df.groupby('category')['revenue'].transform('sum')
           df['key_metrics'] = ((df['revenue'] / df['cat_total_rv']) * 100).round(2)
           
           summ = df[['revenue', 'previous_revenue', 'growth_rate', 'category_top_performer', 'key_metrics']].copy()
           summ['growth_rate'] = summ['growth_rate'].fillna(0)
           
           print('Summary Statistics')
           print(summ) 
           return summ

        except Exception as e:
            print(f"Summary sheet creation failed: {e}")

    def create_pivot_tables(df):
        try:
            df['revenue']  = df['quantity']*df['price']
            piv1 = df.pivot_table(index = 'product', columns = 'region', values = 'revenue', fill_value = 0).round(2)
            piv2 = df.pivot_table(index = 'date', columns = 'product', values = 'revenue', fill_value = 0).round(2)
            piv3 = df.pivot_table(index = 'category', columns = 'region', values = 'revenue', fill_value = 0).round(2)

            print('Revenue by product/region')
            print(piv1)
            print('\nSales by time period')
            print(piv2)
            print('\nPerformance by category')
            print(piv3)

            return piv1, piv2, piv3

        except Exception as e:
            print(f"Pivot tabel creation failed: {e}")

    def create_charts(df):
        try:
            df['revenue'] = df['price'] * df['quantity']

            revenue_trend = df.groupby('date')['revenue'].sum().reset_index()
            category_bn = df.groupby('category')['revenue'].sum().reset_index()
            region_cn = df.groupby('region')['revenue'].sum().reset_index()
            top_product = df.groupby('product')['revenue'].sum().reset_index()

            fig, axes = plt.subplots(2,2,figsize = (12,8))

            revenue_trend.plot(ax = axes[0,0], kind = 'line', x = 'date', y='revenue')
            category_bn.plot(ax = axes[0,1], kind = 'pie', labels = category_bn['category'].to_list(), y = 'revenue',legend = False)
            region_cn.plot(ax = axes[1,0], kind = 'bar', x = 'region', y = 'revenue')
            top_product.plot(ax = axes[1,1], kind = 'barh', x = 'product', y = 'revenue' )

            axes[0,0].set_title('Revenue Trends')
            axes[0,1].set_title('Category Breakdown')
            axes[1,0].set_title('Regional Comparison')
            axes[1,1].set_title('Top Products')

            plt.title('Sales Analytics Summary')
            plt.tight_layout()
            plt.savefig(img)
            plt.show()

            print(f'Figure saved to the system on path: {img}')

            return revenue_trend, category_bn, region_cn, top_product 

        except Exception as e:
            print(f"Chart creation failed: {e}")

    def format_excel_sheets(writer, sheet_name, df):
        try:
            df.to_excel(writer, sheet_name = sheet_name, index = True)
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            header_font = Font(bold = True, color = 'FFFFFFFF', size = 11)
            header_fill = PatternFill(fill_type = 'solid', fgColor = 'FF4472C4')
            header_align = Alignment(horizontal = 'center', vertical = 'center')

            number_format = '#,##0.00'
            percent_format = '0.00%'
            date_format = 'YYYY-MM-DD'

            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            worksheet.row_dimensions[1].height = 25

            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row,
                                   max_col=worksheet.max_column):
                for cell in row:
                    if cell.row % 2 == 0:
                        cell.fill = PatternFill(fill_type='solid', fgColor='DCE6F1')
            
                    if isinstance(cell.value, float):
                        cell.number_format = number_format
                    elif isinstance(cell.value, int):
                        cell.number_format = '#,##0'
                    elif isinstance(cell.value, datetime):
                        cell.number_format = date_format
            
                    cell.alignment = Alignment(horizontal='left', vertical='center')
    
            for column in worksheet.columns:
                max_length = max(
                    len(str(cell.value or '')) for cell in column
                )
                worksheet.column_dimensions[column[0].column_letter].width = max_length + 4
    
            for col_idx, col_name in enumerate(df.columns, start=1):
                if df[col_name].dtype in ['float64', 'int64']:
                    col_letter = worksheet.cell(1, col_idx).column_letter
                    cell_range = f'{col_letter}2:{col_letter}{worksheet.max_row}'
                    worksheet.conditional_formatting.add(
                    cell_range,
                    ColorScaleRule(
                        start_type='min', start_color='FFFF0000', 
                        mid_type='percentile', mid_value=50,
                        mid_color='FFFFFF00',                     
                        end_type='max', end_color='FF00FF00'        
                    )
                    )
            worksheet.freeze_panes  = 'A2'
            worksheet.auto_filter.ref = worksheet.dimensions

        except Exception as e:
            print(f'Excel sheet formatting failed: {e}')

    def create_master_report(df,summ,piv1,piv2,piv3,revenue_trend, category_bn, region_cn, top_product,excel_name):
        try:
            with pd.ExcelWriter(excel_name, engine = 'openpyxl') as writer:
                format_excel_sheets(writer, 'Raw Data', df)
                format_excel_sheets(writer, 'Summary', summ)
                format_excel_sheets(writer, 'ProductxRegion', piv1)
                format_excel_sheets(writer, 'ProductxDate', piv2)
                format_excel_sheets(writer, 'CategoryxRegion', piv3)
                format_excel_sheets(writer, 'Revenue Trend', revenue_trend)
                format_excel_sheets(writer, 'Category breakdown', category_bn)
                format_excel_sheets(writer, 'Regional Comparison', region_cn)
                format_excel_sheets(writer, 'Top Product', top_product)

            print(f'Master report saved to path: {excel_name}')

        except Exception as e:
            print(f"Master report creation failed: {e}")

    summ =None
    p1 = None
    p2 = None
    p3 = None
    rt = None
    cn = None
    rc = None
    tp = None

    data = load_business_data(filename)

    if data is None:
        print("Unable to retreive data")
        return
    while True:
        print("Excel Report Generator")
        print('1. Load Business data')
        print('2. Create Summary Sheet')
        print('3. Generate pivot tables')
        print('4. Create charts')
        print('5. Create master report')
        print('6. Quit')

        ch = input("Enter a choice from 1-6: ")
        if ch == '1':
            load_business_data(filename)
        elif ch == '2':
            summ = create_summary_sheet(data)
        elif ch == '3':
            p1,p2,p3 = create_pivot_tables(data)
        elif ch == '4':
            rt,cn,rc,tp = create_charts(data)
        elif ch == '5':
            create_master_report(data,summ,p1,p2,p3,rt,cn,rc,tp,excel_name)
        elif ch == '6':
            print('Quitting')
            break
        else:
            print('Invalid choice')

excel_gen()